'''
Inserindo dados dados na planilha do Excel
referência: https://www.youtube.com/watch?v=8m4uDS_nyCk

Algoritmo:
- [x] Carregar dados de uma planilha
- [x] Exibir os dados em uma janela Tkinter
- [x] Inserir novos registros na planilha
- [ ] Organizar o código criando funções separadas para:
  - [x] Criar uma função para carregar a janela Tkinter
  - [x] Criar widgets de entrada
  - [x] Configurar o TreeView
  - [x] Configurar os botões
  - [x] Configurar os eventos
  - [x] Criar a função de atualizar m registro
- [ ] Salvar os dados em um banco de dados Sqlite

'''

import tkinter as tk
from tkinter import ttk
from pathlib import Path
import openpyxl


class AppInterface:
    def __init__(self):
        self.ROOT_PATH = Path(__file__).parent
        self.path = self.ROOT_PATH / "data/people.xlsx"
        self.root = tk.Tk()
        self.frame, self.widgets_frame, self.treeFrame = self.criar_frame()
        self.combo_list = ["Subscribed", "Not Subscribed", "Other"]
        self.name_entry, self.age_spinbox, self.status_combobox, self.checkbutton, self.a = self.criar_widgets_entrada()
        self.treeview = self.configurar_treeview()
        self.insert_button, self.update_button, self.mode_switch = self.configurar_botoes()
        self.load_data()

    def criar_frame(self):
        frame = ttk.Frame(self.root)
        frame.pack()
        
        widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
        widgets_frame.grid(row=0, column=0, padx=20, pady=10)
        
        treeFrame = ttk.Frame(frame)
        treeFrame.grid(row=0, column=1, pady=10)
        
        return frame, widgets_frame, treeFrame

    def criar_widgets_entrada(self):
        name_entry = ttk.Entry(self.widgets_frame)
        name_entry.insert(0, "Name")
        name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
        name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

        age_spinbox = ttk.Spinbox(self.widgets_frame, from_=18, to=100)
        age_spinbox.insert(0, "Age")
        age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        
        status_combobox = ttk.Combobox(self.widgets_frame, values=self.combo_list)
        status_combobox.current(0)
        status_combobox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        a = tk.BooleanVar()
        checkbutton = ttk.Checkbutton(self.widgets_frame, text="Employed", variable=a)
        checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")
        
        return name_entry, age_spinbox, status_combobox, checkbutton, a

    def configurar_treeview(self):
        treeScroll = ttk.Scrollbar(self.treeFrame)
        treeScroll.pack(side="right", fill="y")

        cols = ("Name", "Age", "Subscription", "Employment")
        treeview = ttk.Treeview(self.treeFrame, show="headings",
                               yscrollcommand=treeScroll.set, columns=cols, height=13)
        
        treeview.column("Name", width=100)
        treeview.column("Age", width=50)
        treeview.column("Subscription", width=100)
        treeview.column("Employment", width=100)
        treeview.pack()
        
        treeScroll.config(command=treeview.yview)
        treeview.bind('<<TreeviewSelect>>', self.on_treeview_select)
        return treeview

    def configurar_botoes(self):
        button_frame = ttk.Frame(self.widgets_frame)
        button_frame.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")
        
        insert_button = ttk.Button(button_frame, text="Insert", command=self.insert_row)
        insert_button.pack(side="left", padx=5)
        
        update_button = ttk.Button(button_frame, text="Update", command=self.update_row)
        update_button.pack(side="left", padx=5)
        
        separator = ttk.Separator(self.widgets_frame)
        separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")

        mode_switch = ttk.Checkbutton(self.widgets_frame, text="Mode", command=self.toggle_mode)
        mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")
        
        return insert_button, update_button, mode_switch

    def insert_row(self):
        name = self.name_entry.get()
        age = int(self.age_spinbox.get())
        subscription_status = self.status_combobox.get()
        employment_status = "Employed" if self.a.get() else "Unemployed"

        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        row_values = [name, age, subscription_status, employment_status]
        sheet.append(row_values)
        workbook.save(self.path)

        self.treeview.insert('', tk.END, values=row_values)
        
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.status_combobox.set(self.combo_list[0])
        self.checkbutton.state(["!selected"])

    def update_row(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            return
        
        # Pegar valores atuais dos campos
        name = self.name_entry.get()
        age = int(self.age_spinbox.get())
        subscription_status = self.status_combobox.get()
        employment_status = "Employed" if self.a.get() else "Unemployed"
        
        # Atualizar no TreeView
        self.treeview.item(selected_item, values=(name, age, subscription_status, employment_status))
        
        # Atualizar no Excel
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        
        # Encontrar a linha correspondente no Excel
        item_values = self.treeview.item(selected_item)['values']
        row_index = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row == tuple(item_values):
                row_index = idx
                break
        
        if row_index:
            sheet.cell(row=row_index, column=1).value = name
            sheet.cell(row=row_index, column=2).value = age
            sheet.cell(row=row_index, column=3).value = subscription_status
            sheet.cell(row=row_index, column=4).value = employment_status
            workbook.save(self.path)
        
        # Limpar campos após atualização
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.status_combobox.set(self.combo_list[0])
        self.checkbutton.state(["!selected"])

    def on_treeview_select(self, event):
        selected_item = self.treeview.selection()
        if selected_item:
            values = self.treeview.item(selected_item)['values']
            
            # Preencher campos com valores selecionados
            self.name_entry.delete(0, "end")
            self.name_entry.insert(0, values[0])
            
            self.age_spinbox.delete(0, "end")
            self.age_spinbox.insert(0, values[1])
            
            self.status_combobox.set(values[2])
            
            self.checkbutton.state(["selected" if values[3] == "Employed" else "!selected"])

    def toggle_mode(self):
        if self.mode_switch.instate(["selected"]):
            print("Selecionado")
        else:
            print("Desmarcado")

    def load_data(self):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeview.insert('', tk.END, values=value_tuple)

    def run(self):
        self.root.mainloop()

def main():
    app = AppInterface()
    app.run()

if __name__ == "__main__":
    main()
